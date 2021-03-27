import csv
import os 
import shutil
from datetime import date
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Color, PatternFill
from openpyxl.styles import colors
import xlsxwriter


def get_last_row(file, cln):
    result = [0,0]
    #list1 = get_cl(file, "sheet1", cln, 100)
    wb = load_workbook(file)
    ws = wb.worksheets[0]
    range = 100
    range2 = 100
    result1 = 0
    result2 = 0
    while True:
        if ws.cell(row = range, column = 1).internal_value != None:
            if (ws.cell(row = range, column = 1).internal_value) != None:
                result1 =  int(ws.cell(row = range, column = 1).internal_value)
            else:
                result1 = 1
            break
        else:
            range -= 1

    while True: 
        if ws.cell(row = range2, column = 16).internal_value != None:
            if (ws.cell(row = range2, column = 16).internal_value) != None:
                result2 = range2
            else:
                result2 = 1
            break
        else:
            range2 -= 1

    result = [result1, result2]
    wb.close()
    return result
    
def get_list(file):
    liste = []
    with open (file, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            for item in row:
                liste.append(str(item))
    return liste

def get_cl(file, sheet, cln, range):
    list1 = []
    wb =  load_workbook(file)
    ws = wb.worksheets[0]

    list1 = [c.value for c in ws[cln][0:range]]

    wb.close()
    return list1

def get_index(file, mot):
    id = 0
    liste = get_list(file)
    idx = liste.index(mot)
    if idx % 2 == 0:
        id = int(idx/2)
    else: 
        id = int((idx-1)/2)
    return id

def set_jour(test, ind):

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    redFill = PatternFill(start_color='ECEE7C',
                            end_color='ECEE7C',
                            fill_type='solid')
    
    dt = date.today()
    dt_sort = dt.strftime("%d_%b_%Y")

    dest_file = "./journalier/"+dt_sort+".xlsx"
    header = [u'Num', u'Client', u'Référence', u'Gamme', u'Coleur', u'Quantité', u'Défaillance', u'Prix_unitaire', u'Prix_total', u'Remise', u'Net_à_payer', u'Crédit', u'Payé', u'Méthode', u'Temps', u'Copouns']
    try:
        if os.path.isdir("./journalier/"):
            pass
        else:
            os.makedirs("./journalier/")
    except OSError:
        pass

    try:
        if os.path.isfile("./journalier/"+dt_sort+".xlsx"):
            pass
        else:
            wbook = Workbook()
            wbook.create_sheet("Sheet1")
            Sheet1 = wbook.worksheets[0]
            Sheet1.append(header)
            Sheet1.column_dimensions['B'].width = 40
            Sheet1.column_dimensions['C'].width = 25    
            Sheet1.column_dimensions['D'].width = 35
            Sheet1.column_dimensions['E'].width = 30
            Sheet1.column_dimensions['G'].width = 15  
            Sheet1.column_dimensions['H'].width = 15 
            Sheet1.column_dimensions['I'].width = 15 
            Sheet1.column_dimensions['K'].width = 15
            Sheet1.column_dimensions['O'].width = 25
            Sheet1.column_dimensions['P'].width = 30
            for row in Sheet1['A1:P1']:
                for cell in row:
                    cell.border = border
                    cell.fill = redFill
            wbook.save(filename= dest_file)
    except OSError:
        pass
        
    wb =  load_workbook(filename= dest_file)
    ws = wb.worksheets[0]
    ws.append(test)
    ind = ind + 1
    i = str(ind)
    for row in ws['A'+i+':P'+i]:
        for cell in row:
         cell.border = border
    wb.save(dest_file)
    
# print (get_last_row("./journalier/23_Mar_2021.xlsx", 1))

